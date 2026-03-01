"""
extractors.py — Text extraction utilities for multiple file formats.
No domain-specific logic; all analytical logic is delegated to the AI layer.
"""
from __future__ import annotations

import io
import re
from typing import List, Tuple


# ─── Text chunking helpers ────────────────────────────────────────

def chunk_pages(pages: List[str], max_chars: int = 120_000) -> List[str]:
    """
    Merge pages into chunks that fit within the model context window.
    Each chunk contains page markers for traceability.
    """
    chunks: List[str] = []
    current: List[str] = []
    current_len = 0

    for i, page_text in enumerate(pages, start=1):
        marker = f"\n\n--- PAGE {i} ---\n"
        block = marker + page_text.strip()
        block_len = len(block)

        if current_len + block_len > max_chars and current:
            chunks.append("".join(current))
            current = [block]
            current_len = block_len
        else:
            current.append(block)
            current_len += block_len

    if current:
        chunks.append("".join(current))

    return chunks


def extract_raw_text(pages: List[str]) -> str:
    """Return the full document text with page markers for single-pass AI analysis."""
    parts = []
    for i, page_text in enumerate(pages, start=1):
        parts.append(f"\n\n--- PAGE {i} ---\n{page_text.strip()}")
    return "".join(parts)


def guess_title_and_date(pages: List[str]) -> Tuple[str, str]:
    """
    Lightweight heuristic to extract a document title and date from the first page.
    Used only as metadata fallback — the AI extracts the real values.
    """
    title = "Tender Document"
    date = ""

    if not pages:
        return title, date

    lines = [l.strip() for l in pages[0].splitlines() if l.strip()]
    if lines:
        title = lines[0][:150]

    date_pat = re.compile(
        r"\b(\d{1,2}[./]\d{1,2}[./]20\d{2})\b"
        r"|"
        r"\b((?:January|February|March|April|May|June|July|August|September"
        r"|October|November|December|gennaio|febbraio|marzo|aprile|maggio"
        r"|giugno|luglio|agosto|settembre|ottobre|novembre|dicembre"
        r"|Januar|Februar|März|April|Mai|Juni|Juli|August|September"
        r"|Oktober|November|Dezember)\s+20\d{2})\b",
        re.IGNORECASE,
    )

    for line in lines[:15]:
        m = date_pat.search(line)
        if m:
            date = m.group(0)
            break

    return title, date


# ─── Per-format extraction ────────────────────────────────────────

_CHUNK_SIZE = 3_000  # chars per synthetic "page" for non-PDF formats


def _extract_pdf(file_bytes: bytes) -> List[str]:
    import fitz  # PyMuPDF
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    return [doc[i].get_text("text") for i in range(len(doc))]


def _extract_docx(file_bytes: bytes) -> List[str]:
    from docx import Document as DocxDocument
    doc = DocxDocument(io.BytesIO(file_bytes))

    pages: List[str] = []
    current: List[str] = []
    current_len = 0

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        current.append(text)
        current_len += len(text)
        if current_len >= _CHUNK_SIZE:
            pages.append("\n".join(current))
            current = []
            current_len = 0

    if current:
        pages.append("\n".join(current))

    return pages or ["(No text content found in document)"]


def _extract_xlsx(file_bytes: bytes) -> List[str]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    pages: List[str] = []

    for sheet in wb.worksheets:
        rows: List[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            row_text = "\t".join(cells).strip()
            if row_text:
                rows.append(row_text)
        if rows:
            pages.append(f"[Sheet: {sheet.title}]\n" + "\n".join(rows))

    wb.close()
    return pages or ["(No content found in spreadsheet)"]


# ─── Bid-response Excel parser (Knowledge Base) ───────────────────

# Keywords used to detect column roles (case-insensitive substring match)
_REQ_KW    = ("requirement", "requisito", "description", "descrizione",
               "specification", "spec", "question", "domanda", "item",
               "voce", "caratteristica", "feature", "functionality",
               "funzionalità", "testo", "titolo", "title")
_COMPL_KW  = ("compli", "fulfilled", "conform", "soddisf", "status",
               "stato", "y/n", "risposta", "answer", "meet", "met",
               "response", "rispost")
_MAND_KW   = ("mandatory", "obbligat", "required", "m/o", "tipo",
               "type", "priorit", "classe", "class", "category",
               "categoria", "critical", "criticit")
_NOTE_KW   = ("note", "comment", "commento", "osservazione",
               "observation", "detail", "dettaglio", "remark",
               "giustif", "justif", "motiv")

# Values normalised to compliance categories
_VAL_NO      = {"n", "no", "0", "false", "non conforme", "not fulfilled",
                "not compliant", "not met", "non soddisfatto", "x", "no/n"}
_VAL_PARTIAL = {"partial", "parziale", "partially", "p", "part",
                "in parte", "partially compliant", "parzialmente",
                "partially fulfilled"}
_VAL_YES     = {"y", "yes", "si", "sì", "1", "true", "conforme",
                "fulfilled", "compliant", "met", "soddisfatto", "ok"}
_VAL_MAND    = {"m", "mandatory", "obbligatorio", "required", "yes",
                "y", "si", "sì", "1", "true", "critical", "critici",
                "obblig"}


def _col_index(headers: List[str], keywords: tuple) -> int:
    """Return index of first header that contains any keyword (case-insensitive). -1 if none."""
    for i, h in enumerate(headers):
        hl = h.lower()
        if any(kw in hl for kw in keywords):
            return i
    return -1


def _norm(val) -> str:
    """Normalise a cell value to a clean lowercase string."""
    return str(val).strip().lower() if val is not None else ""


def parse_bid_response_excel(file_bytes: bytes, filename: str) -> str:
    """
    Parse a past-bid-response Excel compliance matrix and return a structured
    text summary focused on the information most useful for AI risk analysis:

    - MANDATORY requirements answered N / partially (capability gaps & weaknesses)
    - MANDATORY requirements answered Y (confirmed capabilities)
    - OPTIONAL requirements answered N / partially (lower priority gaps)

    The returned string is meant to be embedded directly in the AI system prompt.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return f"(Could not parse {filename}: {exc})"

    all_sections: List[str] = []

    for sheet in wb.worksheets:
        raw_rows = list(sheet.iter_rows(values_only=True))
        if len(raw_rows) < 2:
            continue

        # Find first non-empty row as header (search up to row 5)
        header_idx = 0
        for i, row in enumerate(raw_rows[:5]):
            if sum(1 for c in row if c is not None) >= 2:
                header_idx = i
                break

        headers = [str(c).strip() if c is not None else "" for c in raw_rows[header_idx]]
        data_rows = raw_rows[header_idx + 1:]
        if not data_rows:
            continue

        req_col   = _col_index(headers, _REQ_KW)
        compl_col = _col_index(headers, _COMPL_KW)
        mand_col  = _col_index(headers, _MAND_KW)
        note_col  = _col_index(headers, _NOTE_KW)

        # Need at least a requirement column and a compliance column to be useful
        if req_col == -1 or compl_col == -1:
            # Fallback: dump first 60 rows as plain text so the AI can still read it
            rows_text = []
            for row in raw_rows[:60]:
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(cells).strip()
                if line.replace("|", "").strip():
                    rows_text.append(line)
            all_sections.append(
                f"[Sheet: {sheet.title}] (columns not auto-detected — raw dump)\n"
                + "\n".join(rows_text)
            )
            continue

        # Bucket rows by compliance + mandatory status
        mand_no:      List[str] = []
        mand_partial: List[str] = []
        mand_yes:     List[str] = []
        opt_no:       List[str] = []
        opt_partial:  List[str] = []

        def _row_line(row, req_col, compl_col, note_col, headers) -> str:
            req_text  = _norm(row[req_col])   if req_col  < len(row) else ""
            note_text = _norm(row[note_col])  if note_col != -1 and note_col < len(row) else ""
            compl_raw = str(row[compl_col]).strip() if compl_col < len(row) and row[compl_col] is not None else ""
            line = f"• [{compl_raw}] {req_text}"
            if note_text and note_text not in ("none", "nan", ""):
                line += f"  → Note: {note_text}"
            return line

        for row in data_rows:
            if all(c is None for c in row):
                continue
            compl_raw  = _norm(row[compl_col]) if compl_col < len(row) else ""
            mand_raw   = _norm(row[mand_col])  if mand_col != -1 and mand_col < len(row) else "m"

            is_mand    = any(kw in mand_raw for kw in _VAL_MAND) if mand_raw else True
            is_no      = compl_raw in _VAL_NO      or any(v == compl_raw for v in _VAL_NO)
            is_partial = compl_raw in _VAL_PARTIAL or any(v in compl_raw for v in ("partial", "parzial", "in parte"))
            is_yes     = compl_raw in _VAL_YES     or any(v == compl_raw for v in _VAL_YES)

            line = _row_line(row, req_col, compl_col, note_col, headers)

            if is_mand:
                if is_no:
                    mand_no.append(line)
                elif is_partial:
                    mand_partial.append(line)
                elif is_yes:
                    mand_yes.append(line)
            else:
                if is_no:
                    opt_no.append(line)
                elif is_partial:
                    opt_partial.append(line)

        parts: List[str] = [f"[Sheet: {sheet.title}]"]

        if mand_no:
            parts.append(
                "MANDATORY — NOT COMPLIANT (N): Inpeco COULD NOT meet these requirements.\n"
                "These reveal CAPABILITY GAPS — flag as HIGH risk if the current tender asks for similar things.\n"
                + "\n".join(mand_no)
            )
        if mand_partial:
            parts.append(
                "MANDATORY — PARTIALLY COMPLIANT: Inpeco only partially met these.\n"
                "These reveal KNOWN WEAKNESSES — flag as MEDIUM-HIGH risk if similar requirements appear.\n"
                + "\n".join(mand_partial)
            )
        if opt_no:
            parts.append(
                "OPTIONAL — NOT COMPLIANT (N): Inpeco could not meet these optional requirements.\n"
                + "\n".join(opt_no)
            )
        if opt_partial:
            parts.append(
                "OPTIONAL — PARTIALLY COMPLIANT:\n"
                + "\n".join(opt_partial)
            )
        if mand_yes:
            # Include confirmed capabilities but keep it compact
            parts.append(
                f"MANDATORY — COMPLIANT (Y): {len(mand_yes)} mandatory requirements met. "
                "Key examples:\n"
                + "\n".join(mand_yes[:20])
                + ("\n  (…and more)" if len(mand_yes) > 20 else "")
            )

        if len(parts) > 1:
            all_sections.append("\n\n".join(parts))

    wb.close()

    if not all_sections:
        return f"(No compliance data found in {filename})"

    return "\n\n".join(all_sections)


def _extract_text(file_bytes: bytes) -> List[str]:
    text = file_bytes.decode("utf-8", errors="replace")
    return [text[i:i + _CHUNK_SIZE] for i in range(0, len(text), _CHUNK_SIZE)] or [""]


# ─── Public dispatcher ────────────────────────────────────────────

SUPPORTED_EXTENSIONS = {"pdf", "docx", "xlsx", "xls", "txt", "csv", "tsv", "md"}


def extract_from_file(file_bytes: bytes, filename: str) -> List[str]:
    """
    Extract text from any supported file type.
    Returns a list of page/chunk strings.
    Never raises — on any error returns a single-item list with a warning message.

    Supported: .pdf, .docx, .xlsx, .xls, .txt, .csv, .tsv, .md
    """
    import zipfile

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    try:
        if ext == "pdf":
            return _extract_pdf(file_bytes)

        elif ext == "docx":
            try:
                return _extract_docx(file_bytes)
            except (zipfile.BadZipFile, Exception):
                # Fallback: file may be an old .doc binary or corrupted —
                # try to salvage whatever readable text is present.
                try:
                    return _extract_text(file_bytes)
                except Exception:
                    return [f"(Could not extract text from {filename}: not a valid DOCX/ZIP file)"]

        elif ext in ("xlsx", "xls"):
            try:
                return _extract_xlsx(file_bytes)
            except (zipfile.BadZipFile, Exception):
                try:
                    return _extract_text(file_bytes)
                except Exception:
                    return [f"(Could not extract text from {filename}: not a valid spreadsheet file)"]

        elif ext in ("txt", "csv", "tsv", "md"):
            return _extract_text(file_bytes)

        else:
            try:
                return _extract_text(file_bytes)
            except Exception:
                return [f"(Could not extract text from: {filename})"]

    except Exception as exc:
        return [f"(Unexpected error reading {filename}: {exc})"]
